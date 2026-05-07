"""REST API views for the Connectivity Matrix Diagram app."""

from io import BytesIO

from django.db import models, transaction
from django.http import HttpResponse
from django.utils.text import slugify
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from nautobot.apps.api import NautobotModelViewSet
from nautobot.dcim.models import Device, Interface

from ..models import ConnectionPlan, ConnectionPlanBatch
from ..drawio import generate_drawio_xml
from ..services.interface_options import EXCLUDED_INTERFACE_TYPES, collect_reserved_interface_ids
from ..stack_plan import generate_stack_plan_template, import_stacks_from_rows, import_stacks_from_xlsx
from .serializers import (
    ConnectionPlanSerializer,
    ConnectionPlanBatchSerializer,
    ConnectionPlanGridSerializer,
    AvailableInterfaceSerializer,
    BatchActionResultSerializer,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


class ConnectionPlanBatchViewSet(NautobotModelViewSet):
    """ViewSet for ConnectionPlanBatch model."""

    queryset = ConnectionPlanBatch.objects.all()
    serializer_class = ConnectionPlanBatchSerializer
    filterset_fields = ["name", "tenant", "location", "status"]

    @action(detail=True, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request, pk=None):
        """Export this batch's connection plans to an XLSX file."""
        if Workbook is None:
            return Response(
                {"error": "openpyxl is not installed; cannot export XLSX."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        batch = self.get_object()
        from ..services.export import build_matrix_workbook

        safe_name = slugify(batch.name) or "batch"
        filename = f"connectivity_matrix_{safe_name}_{batch.pk}.xlsx"
        response = HttpResponse(
            build_matrix_workbook(batch),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"], url_path="import-xlsx")
    def import_xlsx(self, request, pk=None):
        """Import connection plans into this batch from an XLSX file."""
        if load_workbook is None:
            return Response(
                {"error": "openpyxl is not installed; cannot import XLSX."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        batch = self.get_object()
        uploaded = request.FILES.get("file") or request.FILES.get("matrix_file") or request.FILES.get("xlsx")
        if not uploaded:
            return Response(
                {"error": "Missing upload file (expected form field 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        replace = str(request.data.get("replace", "")).lower() in {"1", "true", "yes", "on"}
        if replace:
            batch.connection_plans.all().delete()

        wb = load_workbook(filename=uploaded, data_only=True)
        if "matrix" in wb.sheetnames:
            ws = wb["matrix"]
        elif "Connectivity Matrix" in wb.sheetnames:
            ws = wb["Connectivity Matrix"]
        else:
            ws = wb.active

        # Locate header row and build a case-insensitive column map.
        header_row_idx = None
        column_map = {}
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if not row:
                continue
            normalized = [str(v).strip().lower() if v is not None else "" for v in row]
            if "device a" in normalized and "device b" in normalized:
                header_row_idx = idx
                for col_idx, name in enumerate(normalized):
                    if name:
                        column_map[name] = col_idx
                break

        if header_row_idx is None:
            return Response(
                {"error": "Could not find expected header row (must include 'Device A' and 'Device B')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _get(row, key, default=""):
            col = column_map.get(key)
            if col is None or col >= len(row):
                return default
            value = row[col]
            return "" if value is None else str(value).strip()

        created = 0
        skipped = 0
        errors = []

        next_row_order = (batch.connection_plans.aggregate(models.Max("row_order")).get("row_order__max") or 0) + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            device_a_name = _get(row, "device a")
            device_b_name = _get(row, "device b")
            if not any([device_a_name, device_b_name, _get(row, "device a interface"), _get(row, "device b interface")]):
                skipped += 1
                continue

            interface_a_name = _get(row, "device a interface")
            interface_b_name = _get(row, "device b interface")
            medium = _get(row, "medium") or "RJ45"
            speed = _get(row, "speed") or "1G"
            sfp_a = _get(row, "device a sfp")
            sfp_b = _get(row, "device b sfp")
            notes = _get(row, "notes")

            try:
                device_a = Device.objects.filter(name=device_a_name).first() if device_a_name else None
                device_b = Device.objects.filter(name=device_b_name).first() if device_b_name else None

                interface_a = (
                    Interface.objects.filter(device=device_a, name=interface_a_name).first()
                    if (device_a and interface_a_name)
                    else None
                )
                interface_b = (
                    Interface.objects.filter(device=device_b, name=interface_b_name).first()
                    if (device_b and interface_b_name)
                    else None
                )

                ConnectionPlan.objects.create(
                    batch=batch,
                    row_order=next_row_order,
                    device_a=device_a,
                    device_a_name="" if device_a else device_a_name,
                    interface_a=interface_a,
                    interface_a_name="" if interface_a else interface_a_name,
                    sfp_a=sfp_a,
                    medium=medium,
                    speed=speed,
                    device_b=device_b,
                    device_b_name="" if device_b else device_b_name,
                    interface_b=interface_b,
                    interface_b_name="" if interface_b else interface_b_name,
                    sfp_b=sfp_b,
                    notes=notes,
                    status="draft",
                )
                next_row_order += 1
                created += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({"row": row_idx, "error": str(exc)})

        return Response(
            {
                "replace": replace,
                "created_count": created,
                "skipped_count": skipped,
                "error_count": len(errors),
                "errors": errors[:50],
            }
        )

    @action(detail=True, methods=["post"], url_path="reorder")
    def reorder(self, request, pk=None):
        """
        Update row ordering for connection plans in this batch.

        Expects JSON: {"ordered_ids": ["<plan_uuid>", ...]}
        """
        batch = self.get_object()
        ordered_ids = request.data.get("ordered_ids") or request.data.get("ids")

        if not isinstance(ordered_ids, list):
            return Response(
                {"error": "ordered_ids must be a JSON list of plan IDs."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order_map = {}
        for raw_id in ordered_ids:
            if raw_id in order_map:
                continue
            order_map[raw_id] = len(order_map) + 1

        selected = list(batch.connection_plans.filter(pk__in=list(order_map.keys())))
        remaining = list(
            batch.connection_plans.exclude(pk__in=list(order_map.keys())).order_by("row_order", "created", "pk")
        )

        next_order = len(order_map) + 1
        for plan in remaining:
            order_map[str(plan.pk)] = next_order
            next_order += 1
            selected.append(plan)

        for plan in selected:
            plan.row_order = order_map.get(str(plan.pk), plan.row_order)

        with transaction.atomic():
            ConnectionPlan.objects.bulk_update(selected, ["row_order"])

        return Response({"updated_count": len(selected)})

    @action(detail=True, methods=["get"], url_path="export-drawio")
    def export_drawio(self, request, pk=None):
        """Export this batch's connection plans as a draw.io diagram."""
        batch = self.get_object()

        connections = []
        for idx, plan in enumerate(batch.connection_plans.all().order_by("row_order", "created", "pk"), start=1):
            device_a = plan.device_a.name if plan.device_a else (plan.device_a_name or "")
            device_b = plan.device_b.name if plan.device_b else (plan.device_b_name or "")
            if not device_a or not device_b:
                continue
            connections.append(
                {
                    "row": idx,
                    "device_a": device_a,
                    "interface_a": plan.interface_a.name if plan.interface_a else (plan.interface_a_name or ""),
                    "medium": plan.medium or "",
                    "speed": plan.speed or "",
                    "device_b": device_b,
                    "interface_b": plan.interface_b.name if plan.interface_b else (plan.interface_b_name or ""),
                }
            )

        try:
            xml = generate_drawio_xml(connections)
        except Exception as exc:  # noqa: BLE001
            return Response({"error": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        safe_name = slugify(batch.name) or "batch"
        filename = f"connectivity_diagram_{safe_name}_{batch.pk}.drawio"
        response = HttpResponse(xml.encode("utf-8"), content_type="application/xml")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"])
    def validate(self, request, pk=None):
        """
        Validate all connection plans in this batch.

        POST /api/plugins/connectivity-matrix/batches/{id}/validate/
        """
        batch = self.get_object()
        plans = batch.connection_plans.filter(status__in=["draft", "conflict"])

        results = {
            "success_count": 0,
            "error_count": 0,
            "errors": [],
        }

        for plan in plans:
            errors = plan.validate_connection()
            plan.save()

            if errors:
                results["error_count"] += 1
                results["errors"].append({
                    "plan_id": str(plan.id),
                    "errors": errors,
                })
            else:
                results["success_count"] += 1

        return Response(BatchActionResultSerializer(results).data)

    @action(detail=True, methods=["post"])
    def execute(self, request, pk=None):
        """
        Execute all approved connection plans in this batch (create cables).

        POST /api/plugins/connectivity-matrix/batches/{id}/execute/
        """
        batch = self.get_object()
        plans = batch.connection_plans.filter(status="approved")

        results = {
            "success_count": 0,
            "error_count": 0,
            "errors": [],
        }

        for plan in plans:
            try:
                plan.execute()
                results["success_count"] += 1
            except Exception as e:
                results["error_count"] += 1
                results["errors"].append({
                    "plan_id": str(plan.id),
                    "error": str(e),
                })

        # Update batch status
        if results["error_count"] == 0 and results["success_count"] > 0:
            batch.status = "completed"
        elif results["success_count"] > 0:
            batch.status = "partial"
        batch.save()

        return Response(BatchActionResultSerializer(results).data)

    @action(detail=True, methods=["post"])
    def approve_all(self, request, pk=None):
        """
        Approve all validated connection plans in this batch.

        POST /api/plugins/connectivity-matrix/batches/{id}/approve_all/
        """
        batch = self.get_object()
        count = batch.connection_plans.filter(status="validated").update(status="approved")
        batch.status = "approved"
        batch.save()

        return Response({"approved_count": count})

    @action(detail=True, methods=["post"], url_path="materialize-missing-devices")
    def materialize_missing_devices(self, request, pk=None):
        """Create unresolved devices/interfaces from typed matrix values."""
        from ..services.materialization import materialize_missing_devices

        batch = self.get_object()
        result = materialize_missing_devices(batch)
        status_code = status.HTTP_400_BAD_REQUEST if result.errors else status.HTTP_200_OK
        return Response(
            {
                "created_devices": result.created_devices,
                "reused_devices": result.reused_devices,
                "created_interfaces": result.created_interfaces,
                "errors": result.errors,
            },
            status=status_code,
        )


class ConnectionPlanViewSet(NautobotModelViewSet):
    """ViewSet for ConnectionPlan model."""

    queryset = ConnectionPlan.objects.select_related(
        "batch",
        "device_a",
        "device_b",
        "interface_a",
        "interface_b",
        "created_cable",
    )
    serializer_class = ConnectionPlanSerializer
    filterset_fields = ["batch", "status", "medium", "speed"]

    @action(detail=True, methods=["patch"], url_path="grid")
    def grid_update(self, request, pk=None):
        """Patch a single row using the Tabulator/grid serializer."""
        plan = self.get_object()
        serializer = ConnectionPlanGridSerializer(plan, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def grid(self, request):
        """
        Get connection plans formatted for Tabulator grid.

        GET /api/plugins/connectivity-matrix/plans/grid/?batch={batch_id}
        """
        batch_id = request.query_params.get("batch")
        if not batch_id:
            return Response(
                {"error": "batch parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = self.queryset.filter(batch_id=batch_id).order_by("row_order", "created", "pk")
        serializer = ConnectionPlanGridSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def validate(self, request, pk=None):
        """
        Validate a single connection plan.

        POST /api/plugins/connectivity-matrix/plans/{id}/validate/
        """
        plan = self.get_object()
        errors = plan.validate_connection()
        plan.save()

        return Response({
            "status": plan.status,
            "errors": errors,
        })

    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        """
        Bulk create connection plans for a batch.

        POST /api/plugins/connectivity-matrix/plans/bulk_create/
        {
            "batch_id": "...",
            "rows": [
                {"device_a_name": "...", "interface_a_name": "...", ...},
                ...
            ]
        }
        """
        batch_id = request.data.get("batch_id")
        rows = request.data.get("rows", [])

        if not batch_id:
            return Response(
                {"error": "batch_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            batch = ConnectionPlanBatch.objects.get(id=batch_id)
        except ConnectionPlanBatch.DoesNotExist:
            return Response(
                {"error": "Batch not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        created = []
        errors = []

        next_row_order = (batch.connection_plans.aggregate(models.Max("row_order")).get("row_order__max") or 0) + 1

        for idx, row_data in enumerate(rows):
            try:
                plan = ConnectionPlan.objects.create(
                    batch=batch,
                    row_order=next_row_order,
                    device_a_name=row_data.get("device_a_name", ""),
                    interface_a_name=row_data.get("interface_a_name", ""),
                    sfp_a=row_data.get("sfp_a", ""),
                    device_b_name=row_data.get("device_b_name", ""),
                    interface_b_name=row_data.get("interface_b_name", ""),
                    sfp_b=row_data.get("sfp_b", ""),
                    medium=row_data.get("medium", "RJ45"),
                    speed=row_data.get("speed", "1G"),
                    row_color=row_data.get("row_color", ""),
                    notes=row_data.get("notes", ""),
                )
                next_row_order += 1
                created.append(str(plan.id))
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        return Response({
            "created_count": len(created),
            "created_ids": created,
            "errors": errors,
        })

    @action(detail=True, methods=["post"])
    def swap(self, request, pk=None):
        """Swap Device A/B endpoint fields for one row."""
        plan = self.get_object()
        plan.swap_endpoints()
        plan.save()
        return Response(ConnectionPlanGridSerializer(plan).data)

    @action(detail=True, methods=["post"])
    def color(self, request, pk=None):
        """Apply a visual color marker to one row."""
        from ..services.row_actions import normalize_row_color

        try:
            color = normalize_row_color(request.data.get("row_color"))
        except ValueError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        plan = self.get_object()
        plan.row_color = color
        plan.save(update_fields=["row_color", "last_updated"])
        return Response(ConnectionPlanGridSerializer(plan).data)

    @action(detail=False, methods=["post"])
    def bulk_swap(self, request):
        """Swap endpoint fields for selected rows."""
        ids = request.data.get("ids", [])
        plans = list(self.queryset.filter(id__in=ids))
        for plan in plans:
            plan.swap_endpoints()
            plan.save()
        return Response({"updated_count": len(plans)})

    @action(detail=False, methods=["post"])
    def bulk_color(self, request):
        """Apply a visual color marker to selected rows."""
        from ..services.row_actions import normalize_row_color

        try:
            color = normalize_row_color(request.data.get("row_color"))
        except ValueError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        ids = request.data.get("ids", [])
        count = self.queryset.filter(id__in=ids).update(row_color=color)
        return Response({"updated_count": count, "row_color": color})


class AvailableInterfacesView(APIView):
    """
    API endpoint for getting available (uncabled) interfaces for a device.

    Used by the Tabulator grid for dynamic dropdown population.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        GET /api/plugins/connectivity-matrix/available-interfaces/?device_id={id}

        Returns interfaces that:
        - Belong to the specified device
        - Do not have a cable attached
        - Do not have a pending connection plan
        - Are physical interfaces (not virtual, lag, bridge)
        """
        device_id = request.query_params.get("device_id")
        plan_id = request.query_params.get("plan_id")
        batch_id = request.query_params.get("batch_id") or request.query_params.get("batch")
        if not device_id:
            return Response(
                {"error": "device_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            device = Device.objects.get(id=device_id)
        except Device.DoesNotExist:
            return Response(
                {"error": "Device not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get interfaces without cables
        interfaces = Interface.objects.filter(
            device=device,
            cable__isnull=True,
        ).exclude(
            # Exclude virtual interface types
            type__in=EXCLUDED_INTERFACE_TYPES
        )

        # Exclude interfaces with pending connection plans (except for the current plan if provided).
        plan_rows = ConnectionPlan.objects.filter(
            models.Q(interface_a__device=device) | models.Q(interface_b__device=device)
        )
        if batch_id:
            plan_rows = plan_rows.filter(batch_id=batch_id)
        if plan_id:
            plan_rows = plan_rows.exclude(pk=plan_id)
        plan_rows = plan_rows.values("status", "interface_a_id", "interface_b_id")
        pending_interface_ids = collect_reserved_interface_ids(plan_rows)

        # Exclude pending interfaces
        interfaces = interfaces.exclude(id__in=pending_interface_ids)

        # Order by name
        interfaces = interfaces.order_by("name")

        # Serialize for dropdown
        data = [
            {"value": str(iface.id), "label": iface.name, "type": iface.type}
            for iface in interfaces
        ]

        return Response(data)


class AvailableDevicesView(APIView):
    """
    API endpoint for getting available devices for the dropdown.

    Supports filtering by tenant and location.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        GET /api/plugins/connectivity-matrix/available-devices/?tenant={id}&location={id}

        Returns devices filtered by tenant and/or location.
        """
        def _parse_multi(key):
            values = request.query_params.getlist(key)
            if values:
                out = []
                for val in values:
                    if val is None:
                        continue
                    for part in str(val).split(","):
                        part = part.strip()
                        if part:
                            out.append(part)
                return out
            value = request.query_params.get(key)
            if not value:
                return []
            return [v.strip() for v in str(value).split(",") if v.strip()]

        queryset = Device.objects.all()

        tenant_id = request.query_params.get("tenant")
        if tenant_id:
            queryset = queryset.filter(tenant_id=tenant_id)

        location_id = request.query_params.get("location")
        if location_id:
            queryset = queryset.filter(location_id=location_id)

        status_ids = _parse_multi("status_id")
        if status_ids:
            queryset = queryset.filter(status_id__in=status_ids)

        role_ids = _parse_multi("role_id")
        if role_ids:
            queryset = queryset.filter(role_id__in=role_ids)

        exclude_role_ids = _parse_multi("exclude_role_id")
        if exclude_role_ids:
            queryset = queryset.exclude(role_id__in=exclude_role_ids)

        # Limit results for performance
        limit = int(request.query_params.get("limit", 500))
        queryset = queryset.order_by("name")[:limit]

        data = [
            {"value": str(dev.id), "label": dev.name}
            for dev in queryset
        ]

        return Response(data)


class StackPlanTemplateView(APIView):
    """Download an XLSX template for stack-plan imports."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            content = generate_stack_plan_template()
        except Exception as exc:  # noqa: BLE001
            return Response({"error": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="stack_plan_template.xlsx"'
        return response


class StackPlanImportView(APIView):
    """Import stack plan XLSX and create planned stack devices."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        uploaded = request.FILES.get("file") or request.FILES.get("stack_plan_file") or request.FILES.get("xlsx")
        if not uploaded:
            return Response(
                {"error": "Missing upload file (expected form field 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            created, skipped, error_rows, errors = import_stacks_from_xlsx(uploaded.read())
        except Exception as exc:  # noqa: BLE001
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "created_devices": created,
                "skipped_rows": skipped,
                "error_rows": error_rows,
                "errors": errors[:50],
            }
        )


class StackPlanMaterializeView(APIView):
    """Materialize stack/device rows submitted by the online stack builder."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        rows = request.data.get("rows") or []
        defaults = request.data.get("defaults") or {}
        if not isinstance(rows, list):
            return Response({"error": "Expected 'rows' to be a list."}, status=status.HTTP_400_BAD_REQUEST)
        if not rows:
            return Response({"error": "Add at least one stack/member row before materializing."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            created, skipped, error_rows, errors = import_stacks_from_rows(rows, defaults=defaults)
        except Exception as exc:  # noqa: BLE001
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "created_devices": created,
                "skipped_rows": skipped,
                "error_rows": error_rows,
                "errors": errors[:100],
            }
        )
