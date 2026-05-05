"""Excel export for the online connectivity matrix."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADERS = [
    "Device A",
    "Device A Interface",
    "Device A SFP",
    "Medium",
    "Speed",
    "Device B",
    "Device B Interface",
    "Device B SFP",
    "Status",
    "Notes",
]


def build_matrix_workbook(batch) -> bytes:
    """Return an XLSX workbook containing the current online matrix."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "matrix"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    for plan in batch.connection_plans.order_by("row_order", "created", "pk").all():
        sheet.append(
            [
                plan.device_a_display,
                plan.interface_a_display,
                plan.sfp_a,
                plan.medium,
                plan.speed,
                plan.device_b_display,
                plan.interface_b_display,
                plan.sfp_b,
                getattr(plan, "status", ""),
                plan.notes,
            ]
        )
        if getattr(plan, "row_color", ""):
            color = plan.row_color.replace("#", "").upper()
            for cell in sheet[sheet.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 40)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
