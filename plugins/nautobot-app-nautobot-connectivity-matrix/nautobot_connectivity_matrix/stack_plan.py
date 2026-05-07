"""Stack-plan import/template helpers.

This is derived from the legacy Job implementation:
`nautobot_jobs_repo/jobs/connectivity_matrix_diagram/connectivity_matrix_diagram.py`.
"""

from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

from django.db import transaction
from django.contrib.contenttypes.models import ContentType

from nautobot.dcim.models import Device, DeviceType, Location, Module, ModuleBay, ModuleType, Platform
from nautobot.extras.models import Role, Status
from nautobot.tenancy.models import Tenant
from nautobot_connectivity_matrix.services.interface_finalization import (
    ensure_device_type_interfaces,
    normalize_device_interface_types,
    populate_module_interfaces_for_device,
    set_device_interface_status,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None
    DataValidation = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


STACK_PLAN_HEADERS = [
    "stack_id",
    "stack_hostname",
    "member_position",
    "device_type",
    "module_type",
    "tenant",
    "location",
    "role",
    "platform",
    "status",
]


def _model_label_from_value(model, value: str, *, label_field: str = "name") -> str:
    """Resolve a UI/API value that may be a UUID, slug, name, model, or part number to an import label."""
    key = str(value or "").strip()
    if not key:
        return ""

    lookups = [{"pk": key}, {f"{label_field}__iexact": key}]
    if any(field.name == "slug" for field in model._meta.fields):
        lookups.append({"slug__iexact": key})
    if model in (DeviceType, ModuleType):
        lookups.extend([{"model__iexact": key}, {"part_number__iexact": key}])

    for lookup in lookups:
        try:
            obj = model.objects.filter(**lookup).first()
        except Exception:  # noqa: BLE001
            obj = None
        if obj:
            if model in (DeviceType, ModuleType):
                return obj.model
            return getattr(obj, label_field, str(obj))
    return key


def import_stacks_from_rows(rows: Iterable[Dict[str, str]], defaults: Optional[Dict[str, str]] = None) -> Tuple[int, int, int, List[Dict[str, str]]]:
    """Import stack/device rows submitted by the online stack builder."""
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; cannot import stack plan.")

    defaults = defaults or {}
    default_status_label = _model_label_from_value(Status, defaults.get("status") or "Planned") or "Planned"
    default_tenant_label = _model_label_from_value(Tenant, defaults.get("tenant"))
    default_location_label = _model_label_from_value(Location, defaults.get("location"))
    default_role_label = _model_label_from_value(Role, defaults.get("role"))
    default_platform_label = _model_label_from_value(Platform, defaults.get("platform"))
    default_device_type_label = _model_label_from_value(DeviceType, defaults.get("device_type"), label_field="model")

    wb = Workbook()
    ws = wb.active
    ws.title = "stack_plan"
    ws.append(STACK_PLAN_HEADERS)

    for idx, row in enumerate(rows, start=1):
        stack_id = str(row.get("stack_id") or row.get("stack_key") or idx).strip()
        member_position = str(row.get("member_position") or "1").strip()
        ws.append(
            [
                stack_id,
                str(row.get("stack_hostname") or "").strip(),
                member_position,
                _model_label_from_value(DeviceType, row.get("device_type") or default_device_type_label, label_field="model"),
                _model_label_from_value(ModuleType, row.get("module_type"), label_field="model"),
                _model_label_from_value(Tenant, row.get("tenant") or default_tenant_label),
                _model_label_from_value(Location, row.get("location") or default_location_label),
                _model_label_from_value(Role, row.get("role") or default_role_label),
                _model_label_from_value(Platform, row.get("platform") or default_platform_label),
                _model_label_from_value(Status, row.get("status") or default_status_label),
            ]
        )

    out = BytesIO()
    wb.save(out)
    return import_stacks_from_xlsx(out.getvalue(), default_status_name=default_status_label)


def generate_stack_plan_template(*, max_rows: int = 2000) -> bytes:
    """Generate an XLSX template for stack plan imports."""
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; cannot generate stack-plan template.")

    wb = Workbook()
    ws = wb.active
    ws.title = "stack_plan"
    ws.freeze_panes = "A2"

    ws.append(STACK_PLAN_HEADERS)
    for col_idx in range(1, len(STACK_PLAN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    ws.auto_filter.ref = f"A1:{get_column_letter(len(STACK_PLAN_HEADERS))}1"

    header_fill = PatternFill("solid", fgColor="FF4F81BD")
    header_font = Font(bold=True, color="FFFFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data validation drop-down lists for key columns, using a hidden sheet as the source.
    if DataValidation is not None:
        lists = wb.create_sheet("_lists")
        lists.sheet_state = "hidden"

        device_ct = ContentType.objects.get(app_label="dcim", model="device")

        statuses = list(
            Status.objects.filter(content_types=device_ct)
            .order_by("name")
            .values_list("name", flat=True)
        )
        roles = list(
            Role.objects.filter(content_types=device_ct)
            .order_by("name")
            .values_list("name", flat=True)
        )
        tenants = list(Tenant.objects.order_by("name").values_list("name", flat=True))
        locations = list(Location.objects.order_by("name").values_list("name", flat=True))
        platforms = list(Platform.objects.order_by("name").values_list("name", flat=True))
        device_types = list(DeviceType.objects.order_by("model").values_list("model", flat=True))
        module_types = list(ModuleType.objects.order_by("model").values_list("model", flat=True))

        def _write_list(col_idx: int, values: List[str]) -> str:
            col_letter = get_column_letter(col_idx)
            for row_idx, value in enumerate(values, start=1):
                lists[f"{col_letter}{row_idx}"] = value
            if not values:
                return ""
            return f"={lists.title}!${col_letter}$1:${col_letter}${len(values)}"

        # One column per list on the hidden sheet.
        range_status = _write_list(1, statuses)
        range_role = _write_list(2, roles)
        range_tenant = _write_list(3, tenants)
        range_location = _write_list(4, locations)
        range_platform = _write_list(5, platforms)
        range_device_type = _write_list(6, device_types)
        range_module_type = _write_list(7, module_types)

        max_row = max_rows + 1  # header is row 1

        def _add_list_validation(col_letter: str, formula: str, *, allow_blank: bool = True):
            if not formula:
                return
            dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
            dv.errorTitle = "Invalid value"
            dv.error = "Please select a value from the drop-down list."
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{max_row}")

        # Column letters correspond to STACK_PLAN_HEADERS order.
        _add_list_validation("D", range_device_type)  # device_type
        _add_list_validation("E", range_module_type)  # module_type
        _add_list_validation("F", range_tenant)  # tenant
        _add_list_validation("G", range_location)  # location
        _add_list_validation("H", range_role)  # role
        _add_list_validation("I", range_platform)  # platform
        _add_list_validation("J", range_status)  # status

        # Basic numeric validation for stack_id and member_position.
        dv_int = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
        dv_int.errorTitle = "Invalid number"
        dv_int.error = "Enter a whole number >= 1."
        ws.add_data_validation(dv_int)
        dv_int.add(f"A2:A{max_row}")
        dv_int.add(f"C2:C{max_row}")

    # Provide a couple of empty rows for convenience.
    for _ in range(2, min(max_rows + 2, 50)):
        ws.append([""] * len(STACK_PLAN_HEADERS))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def import_stacks_from_xlsx(content: bytes, *, default_status_name: str = "Planned") -> Tuple[int, int, int, List[Dict[str, str]]]:
    """Import stacks from an XLSX stack plan.

    Returns: (created_devices, skipped_rows, error_rows, errors)
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed; cannot import stack plan.")

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb["stack_plan"] if "stack_plan" in wb.sheetnames else wb.active

    # Build header mapping
    header = [str(v).strip().lower() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: idx for idx, name in enumerate(header) if name}

    missing = [h for h in STACK_PLAN_HEADERS if h not in col]
    if missing:
        raise ValueError(f"Missing required columns in header row: {', '.join(missing)}")

    planned_status = Status.objects.filter(name=default_status_name).first()

    def _cell(row, key):
        value = row[col[key]]
        return "" if value is None else str(value).strip()

    dt_cache: Dict[str, Optional[DeviceType]] = {}
    mt_cache: Dict[str, Optional[ModuleType]] = {}
    tenant_cache: Dict[str, Optional[Tenant]] = {}
    loc_cache: Dict[str, Optional[Location]] = {}
    role_cache: Dict[str, Optional[Role]] = {}
    platform_cache: Dict[str, Optional[Platform]] = {}
    status_cache: Dict[str, Optional[Status]] = {}

    def _has_field(model, field: str) -> bool:
        try:
            model._meta.get_field(field)
            return True
        except Exception:  # noqa: BLE001
            return False

    def _get_dt(value: str) -> Optional[DeviceType]:
        key = (value or "").strip()
        if not key:
            return None
        if key in dt_cache:
            return dt_cache[key]
        dt_cache[key] = (
            DeviceType.objects.filter(model__iexact=key).first()
            or DeviceType.objects.filter(part_number__iexact=key).first()
            or DeviceType.objects.filter(model__icontains=key).first()
            or DeviceType.objects.filter(part_number__icontains=key).first()
        )
        return dt_cache[key]

    def _get_mt(value: str) -> Optional[ModuleType]:
        key = (value or "").strip()
        if not key:
            return None
        if key in mt_cache:
            return mt_cache[key]
        mt_cache[key] = (
            ModuleType.objects.filter(model__iexact=key).first()
            or ModuleType.objects.filter(part_number__iexact=key).first()
            or ModuleType.objects.filter(model__icontains=key).first()
            or ModuleType.objects.filter(part_number__icontains=key).first()
        )
        return mt_cache[key]

    def _get_tenant(value: str) -> Optional[Tenant]:
        key = (value or "").strip()
        if not key:
            return None
        if key in tenant_cache:
            return tenant_cache[key]
        tenant_q = Tenant.objects.filter(name=key).first()
        if not tenant_q and _has_field(Tenant, "slug"):
            tenant_q = Tenant.objects.filter(slug=key).first()
        tenant_cache[key] = tenant_q
        return tenant_cache[key]

    def _get_loc(value: str) -> Optional[Location]:
        key = (value or "").strip()
        if not key:
            return None
        if key in loc_cache:
            return loc_cache[key]
        loc_q = Location.objects.filter(name=key).first()
        if not loc_q and _has_field(Location, "slug"):
            loc_q = Location.objects.filter(slug=key).first()
        loc_cache[key] = loc_q
        return loc_cache[key]

    def _get_role(value: str) -> Optional[Role]:
        key = (value or "").strip()
        if not key:
            return None
        if key in role_cache:
            return role_cache[key]
        role_q = Role.objects.filter(name=key).first()
        if not role_q and _has_field(Role, "slug"):
            role_q = Role.objects.filter(slug=key).first()
        role_cache[key] = role_q
        return role_cache[key]

    def _get_platform(value: str) -> Optional[Platform]:
        key = (value or "").strip()
        if not key:
            return None
        if key in platform_cache:
            return platform_cache[key]
        platform_q = Platform.objects.filter(name=key).first()
        if not platform_q and _has_field(Platform, "slug"):
            platform_q = Platform.objects.filter(slug=key).first()
        platform_cache[key] = platform_q
        return platform_cache[key]

    def _get_status(value: str) -> Optional[Status]:
        key = (value or "").strip()
        if not key:
            return None
        if key in status_cache:
            return status_cache[key]
        status_cache[key] = Status.objects.filter(name=key).first()
        return status_cache[key]

    def _unique_device_name(base: str) -> str:
        if not Device.objects.filter(name=base).exists():
            return base
        i = 1
        while True:
            candidate = f"{base}-{i}"
            if not Device.objects.filter(name=candidate).exists():
                return candidate
            i += 1

    prefix_order = {"Gi": 0, "Te": 1, "Fi": 2, "Twe": 3, "Fo": 4, "Hu": 5, "Eth": 6, "Fa": 7}

    def _abbr(name: str) -> str:
        mapping = {
            "TenGigabitEthernet": "Te",
            "GigabitEthernet": "Gi",
            "FastEthernet": "Fa",
            "Ethernet": "Eth",
            "HundredGigE": "Hu",
            "FortyGigE": "Fo",
            "TwentyFiveGigE": "Twe",
        }
        for long, short in mapping.items():
            if name.startswith(long):
                return name.replace(long, short, 1)
        return name

    def _sort_key(name: str):
        abbr = _abbr(name)
        parts = re.split(r"([A-Za-z]+)|/", abbr)
        parts = [p for p in parts if p not in ("", None)]
        pref = parts[0] if parts else abbr
        nums = [int(p) for p in parts[1:] if p.isdigit()]
        return (*nums, prefix_order.get(pref, 99), abbr)

    stacks: Dict[str, List[Tuple[int, List[str]]]] = defaultdict(list)
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stack_id = _cell(row, "stack_id")
        if not stack_id:
            continue
        stacks[stack_id].append(
            (
                idx,
                [
                    _cell(row, "stack_id"),
                    _cell(row, "stack_hostname"),
                    _cell(row, "member_position"),
                    _cell(row, "device_type"),
                    _cell(row, "module_type"),
                    _cell(row, "tenant"),
                    _cell(row, "location"),
                    _cell(row, "role"),
                    _cell(row, "platform"),
                    _cell(row, "status"),
                ],
            )
        )

    created_devices = 0
    skipped = 0
    error_rows = 0
    errors: List[Dict[str, str]] = []

    with transaction.atomic():
        for stack_id, entries in stacks.items():
            stack_hostname = next((r[1][1] for r in entries if r[1][1]), None)
            if not stack_hostname:
                skipped += len(entries)
                continue

            tenant = next((_get_tenant(r[1][5]) for r in entries if r[1][5]), None)
            location = next((_get_loc(r[1][6]) for r in entries if r[1][6]), None)
            role = next((_get_role(r[1][7]) for r in entries if r[1][7]), None)
            platform = next((_get_platform(r[1][8]) for r in entries if r[1][8]), None)
            status = next((_get_status(r[1][9]) for r in entries if r[1][9]), planned_status)

            existing = Device.objects.filter(name=stack_hostname, tenant=tenant, location=location).first()
            if existing and existing.status and existing.status.name == default_status_name:
                existing.delete()

            device_name = _unique_device_name(stack_hostname)

            first_dt_name = next((r[1][3] for r in entries if r[1][3]), None)
            first_dt = _get_dt(first_dt_name) if first_dt_name else None

            stack_device = Device.objects.create(
                name=device_name,
                device_type=first_dt,
                tenant=tenant,
                location=location,
                role=role,
                platform=platform,
                status=status,
            )
            created_devices += 1

            for row_idx, row_values in sorted(entries, key=lambda x: (x[1][2] or "0")):
                member_position = row_values[2]
                device_type_name = row_values[3]
                module_type_name = row_values[4]

                dt = _get_dt(device_type_name)
                if not dt:
                    error_rows += 1
                    errors.append({"row": str(row_idx), "error": f"DeviceType '{device_type_name}' not found"})
                    continue
                try:
                    pos = str(int(member_position)) if member_position else "1"
                except Exception:
                    pos = "1"

                ensure_device_type_interfaces(
                    stack_device,
                    dt,
                    member_position=pos,
                    status=status,
                )

                if module_type_name:
                    module_names = [m.strip() for m in str(module_type_name).replace(";", ",").split(",") if m.strip()]
                    for mod_name in module_names:
                        mt = _get_mt(mod_name)
                        if not mt:
                            errors.append({"row": str(row_idx), "error": f"ModuleType '{mod_name}' not found"})
                            continue

                        bay_name = f"Network Module {pos}"
                        bay = ModuleBay.objects.filter(parent_device=stack_device, position=pos, name__icontains="Network Module").first()
                        if not bay:
                            bay = ModuleBay.objects.filter(parent_device=stack_device, name=bay_name).first()
                        if not bay and pos == "1":
                            bay = ModuleBay.objects.filter(parent_device=stack_device, name="Network Module").first()
                        if not bay:
                            bay = ModuleBay.objects.create(parent_device=stack_device, name=bay_name, position=pos)
                        elif not bay.position:
                            bay.position = pos
                            bay.save()

                        try:
                            installed_module = bay.installed_module
                        except Module.DoesNotExist:
                            installed_module = None
                        if installed_module and installed_module.module_type == mt:
                            module = installed_module
                            changed = False
                            if status and module.status_id != status.pk:
                                module.status = status
                                changed = True
                            if tenant and module.tenant_id != tenant.pk:
                                module.tenant = tenant
                                changed = True
                            if module.location_id:
                                module.location = None
                                changed = True
                            if changed:
                                module.validated_save()
                        else:
                            if installed_module:
                                installed_module.delete()
                            module = Module(
                                module_type=mt,
                                parent_module_bay=bay,
                                status=status or planned_status,
                                tenant=tenant,
                            )
                            module.validated_save()

            populate_module_interfaces_for_device(stack_device, status=status, prune_stale=True)
            normalize_device_interface_types(stack_device)
            set_device_interface_status(stack_device, status)

    return created_devices, skipped, error_rows, errors
